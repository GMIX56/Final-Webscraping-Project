import openpyxl as xl
from openpyxl.styles import Font

#create a new excel document
wb = xl.Workbook()


ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1,title='Second Sheet')

#write content to a cell
ws['A1'] = 'Invoice'

ws['A1'].font = Font(name='Times New Roman',size=24,italic=False,bold=True)

headerFont = Font(name='Times New Roman',size=24,italic=False,bold=True) 

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws.merge_cells('A1:B1')


#unmerge
#ws.unmerge_cells('A1:B1')


ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = Font(size=16,bold=True)

ws['B8'] = '=SUM(B2:B7)'

#change the column width
ws.column_dimensions['A'].width = 25

wb.save("PythontoExcel.xlsx")


# Read the excel file - 'ProduceReport.xlsx' thtat you created earlier
# write all the contents of this file to 'Second Sheet' in the current
# workbook

# display the Grand Total and Average of 'Amt Sold' and 'Total'
# at the bottom of the list along with appropriate labels

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb('ProduceReport') # activating the sheet that you want to read from

